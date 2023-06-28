import numpy as np
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

import datetime
import openpyxl
import pandas as pd
import os
import shutil

path_asc = './Data/ssq_asc.txt'


def debug_print(message):
    """打印带时间戳的调试信息到控制台
    Args:
        message (字符串): 调试信息
    """

    # 获取当前时间
    current_time = datetime.datetime.now()
    # 格式化时间戳字符串
    timestamp_str = current_time.strftime('[%Y-%m-%d %H:%M:%S]')
    # 打印带时间戳的调试信息
    print(f'[Info]{timestamp_str} {message}')


def get_latest_lottery_results(year, rows):
    """读取开奖该数据

    Args:
        year (int): 开奖年份(yyyy) 或 (yyyy-yyyy)
        rows (int): 开奖数据行

    Returns:
        list: 返回开奖数据前9列
              分别为 ->> 期号、开奖日期、红球1、红球2、红球3、红球4、红球5、红球6、蓝球
    """

    # 筛选期数符合年份的开奖数据
    with open(path_asc, 'r') as f:
        data_lottery = []
        # 指定年份
        if year != None:
            for line in f:
                fields = line.strip().split()[:9]
                # 如果指定年份期间
                if '-' in year:
                    start_year = int(year.split('-')[0])
                    end_year = int(year.split('-')[1])
                    if start_year <= int(fields[0][:4]) <= end_year:
                        data_lottery.append(fields)
                # 不指定年份期间
                else:
                    if fields[0].startswith(year):
                        data_lottery.append(fields)
        # 不指定年份
        else:
            for line in f:
                fields = line.strip().split()[:9]
                data_lottery.append(fields)

    # 返回指定行数的数据
    if rows != None:
        if int(rows) > len(data_lottery):
            debug_print(f'期数不足{rows}期, 只返回{len(data_lottery)}期...')
            rows = len(data_lottery)
        data_lottery = data_lottery[-int(rows):]

    # 将第2列以后的所有数据转换为整数类型
    for i in range(len(data_lottery)):
        data_lottery[i][2:] = [int(x) for x in data_lottery[i][2:]]

    # 返回开奖数据
    return data_lottery


def calc_sum_red_balls(data_lottery):
    """计算红球和值

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据和值列(共1列)
    """

    # 计算红球和值
    debug_print('计算红球和值...')
    for i in range(len(data_lottery)):
        data_lottery[i].append(sum([x for x in data_lottery[i][2:8]]))
    # 返回开奖数据
    return data_lottery

# 计算开奖和值


def calc_sum_lottery(data_lottery):
    """计算开奖和值

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据和值列(共1列)
    """

    # 计算开奖和值
    debug_print('计算开奖(红球和篮球)和值...')
    for i in range(len(data_lottery)):
        data_lottery[i].append(sum([x for x in data_lottery[i][2:9]]))
    # 返回开奖数据
    return data_lottery


def calc_offset_red_balls(data_lottery):
    """计算红球横向差值偏移量

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据红球横向差值偏移量列(共5列)
    """

    # 计算每个红球横向差值偏移量
    debug_print('计算红球横向差值偏移量...')
    for i in range(len(data_lottery)):
        data_lottery[i].append(data_lottery[i][3] - data_lottery[i][2])
        data_lottery[i].append(data_lottery[i][4] - data_lottery[i][3])
        data_lottery[i].append(data_lottery[i][5] - data_lottery[i][4])
        data_lottery[i].append(data_lottery[i][6] - data_lottery[i][5])
        data_lottery[i].append(data_lottery[i][7] - data_lottery[i][6])

    # 返回开奖数据
    return data_lottery


def calc_sum_offset_red_balls(data_lottery):
    """计算红球横向差值偏移量和值

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据红球横向差值偏移量和值列(共1列)
    """

    # 计算红球横向差值偏移量和值
    debug_print('计算红球横向差值偏移量和值...')
    for i in range(len(data_lottery)):
        data_lottery[i].append(sum([x for x in data_lottery[i][11:16]]))
    # 返回开奖数据
    return data_lottery


def calc_vertical_offset_red_balls(data_lottery):
    """计算红球纵向差值偏移量

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据红球纵向差值偏移量列(共6列)
    """

    # 计算每个红球纵向差值偏移量
    debug_print('计算红球纵向差值偏移量...')
    for i in range(len(data_lottery)):
        for j in range(2, 8):
            if i == 0:
                data_lottery[i].append(0)
            else:
                data_lottery[i].append(data_lottery[i][j] - data_lottery[i-1][j])

    # 返回开奖数据
    return data_lottery


def calc_sum_vertical_offset_red_balls(data_lottery):
    """计算红球纵向差值偏移量和值

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据红球纵向差值偏移量和值列(共1列)
    """

    # 计算红球纵向差值偏移量和值
    debug_print('计算红球纵向差值偏移量和值...')
    for i in range(len(data_lottery)):
        data_lottery[i].append(sum([x for x in data_lottery[i][17:23]]))

    # 返回开奖数据
    return data_lottery


def calc_continuous_number_red_balls(data_lottery):
    """计算红球连号出现的次数

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据连号出现的次数列(共1列)
    """

    # 计算连号出现的次数
    debug_print('计算连号出现的次数...')
    num = 0
    for i in range(len(data_lottery)):
        for j in range(2, 7):
            if (data_lottery[i][j] - data_lottery[i][j+1]) == -1:
                num += 1

        data_lottery[i].append(num)
        num = 0

    # 返回开奖数据
    return data_lottery


def find_line_num_in_all_combo(data_lottery):
    """ 查找所在all_combo.txt文件中的行号

    Args:
        data_lottery (_type_): _description_

    Returns:
        _type_: _description_
    """

    # 获取所有号码组合
    with open('./Data/all_combos.txt', mode='r', encoding='utf-8') as f:
        all_combos = [line.strip() for line in f.readlines()]

    # 将数据转换为data frame
    df = pd.DataFrame(all_combos)

    for i in range(len(data_lottery)):
        # 将data_lottery的2到7列转换成字符串以逗号分割
        data_lottery_str = ','.join(str(x) for x in data_lottery[i][2:8])

        # 查找所在行号
        line_num = df[df[0].str.contains(data_lottery_str)].index.tolist()

    debug_print('')


def generate_trend_chart(data_lottery):
    """生成双色球红球1-33和篮球1-16的走势图,并计算每个号码的遗漏值

    Args:
        data_lottery (list): 开奖数据

    Returns:
        list: 追加开奖数据红球1-33和篮球1-16的走势图列(共49列)
    """

    # 生成双色球红球1-33和篮球1-16的走势图
    debug_print('生成双色球红球1-33和篮球1-16的走势图...')
    new_data_lottery = [[] for i in range(len(data_lottery))]
    miss_num = 1
    for i in range(len(data_lottery)):
        for j in range(1, 34):

            if j in data_lottery[i][2:8]:
                new_data_lottery[i].append(j)
                miss_num = 1

            else:
                # 计算历史遗漏值
                if i == 0:
                    new_data_lottery[i].append('y%d' % miss_num)
                elif 'y' in str(new_data_lottery[i-1][j-1]):
                    miss_num = int(new_data_lottery[i-1][j-1].replace('y', '')) + 1
                    new_data_lottery[i].append('y%d' % miss_num)
                else:
                    miss_num = 1
                    new_data_lottery[i].append('y%d' % miss_num)

        miss_num = 1
        for j in range(1, 17):
            if j == data_lottery[i][8]:
                new_data_lottery[i].append(j)
                miss_num = 1
            else:
                # 计算历史遗漏值
                if i == 0:
                    new_data_lottery[i].append('y%d' % miss_num)
                elif 'y' in str(new_data_lottery[i-1][j+32]):
                    miss_num = int(new_data_lottery[i-1][j+32].replace('y', '')) + 1
                    new_data_lottery[i].append('y%d' % miss_num)
                else:
                    miss_num = 1
                    new_data_lottery[i].append('y%d' % miss_num)

    # 将data_lottery的前2列插入到new_data_lottery前面
    new_data_lottery = [data_lottery[i][:2] + new_data_lottery[i] for i in range(len(data_lottery))]

    # 返回开奖数据
    return new_data_lottery


def print_to_excel(sheet, data_lottery):
    """输出结果到Excel

    Args:
        sheet (str): sheet名称
        data (list): 数据源
    """

    debug_print(f'输出{sheet}到文件...')
    headers = []
    if sheet == 'Raw Data':
        headers = ('Issue No.,'
                   'Date of Draw,'
                   'Red01,Red02,Red03,Red04,Red05,Red06,Blue,'
                   'Red Sum,'
                   'All Sum,'
                   'Red01 Offset(Horizontal),'
                   'Red02 Offset(Horizontal),'
                   'Red03 Offset(Horizontal),'
                   'Red04 Offset(Horizontal),'
                   'Red05 Offset(Horizontal),'
                   'Red Offset Sum(Horizontal),'
                   'Red01 Offset(Vertical),'
                   'Red02 Offset(Vertical),'
                   'Red03 Offset(Vertical),'
                   'Red04 Offset(Vertical),'
                   'Red05 Offset(Vertical),'
                   'Red06 Offset(Vertical),'
                   'Red Offset Sum(Vertical),'
                   'Consecutive').split(',')

    elif sheet == 'Trend Chart':
        headers = 'Issue No.,Date of Draw,' + ','.join([str(i).zfill(2) for i in range(1, 34)] + [str(i).zfill(2) for i in range(1, 17)])
        headers = headers.split(',')

    # 创建DataFrame对象
    df = pd.DataFrame(data_lottery, columns=headers)

    # 创建新分析文件
    if not os.path.exists('./Data/output.xlsx'):
        with pd.ExcelWriter('./Data/output.xlsx', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False, header=True)
    # 追加新sheet前remove旧sheet
    else:
        with pd.ExcelWriter('./Data/output.xlsx', mode='a') as writer:
            try:
                writer.book.remove(writer.book[sheet])
            except:
                pass
            df.to_excel(writer, sheet_name=sheet, index=False, header=True)


# 美化output文件
def beautify_output():
    """美化output文件
    """

    debug_print('格式化输出文件...')

    # 打开output文件
    wb = openpyxl.load_workbook('./Data/output.xlsx')

    # 格式化原始数据
    ws = wb['Raw Data']

    # 设置sheet缩放为85%
    ws.sheet_view.zoomScale = 85

    # 设置标题行冻结
    ws.freeze_panes = 'A2'

    # 设置不显示网格线
    ws.sheet_view.showGridLines = False

    # 设置原始数据样式
    for col in ws.iter_cols():
        pass

    # 格式化走势图
    ws = wb['Trend Chart']

    # 设置sheet缩放为85%
    ws.sheet_view.zoomScale = 85

    # 设置标题行冻结
    ws.freeze_panes = 'A2'

    # 设置不显示网格线
    ws.sheet_view.showGridLines = False

    # 设置走势图样式
    for col in ws.iter_cols():
        # 设置期号宽度和对齐方式
        if col[0].column == 1:
            ws.column_dimensions[col[0].column_letter].width = 10
        # 设置开奖日期宽度
        elif col[0].column == 2:
            ws.column_dimensions[col[0].column_letter].width = 15
        # 设置红球篮球列宽
        elif col[0].column > 2:
            ws.column_dimensions[col[0].column_letter].width = 4

        # 设置单元格格式
        for cell in col:
            # 设置所有数据居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置边框
            cell.border = Border(left=Side(border_style='thin', color='BFBFBF'),
                                 right=Side(border_style='thin', color='BFBFBF'),
                                 top=Side(border_style='thin', color='BFBFBF'),
                                 bottom=Side(border_style='thin', color='BFBFBF'))

            # 设置标题行格式
            if cell.row == 1:
                if cell.col_idx <= 35:
                    cell.fill = PatternFill(fill_type='solid', fgColor='f9cc0d')
                    cell.font = Font(name='Consolas', size=11, color='000000', bold=True)
                elif cell.col_idx >= 36 and cell.col_idx <= 51:
                    cell.fill = PatternFill(fill_type='solid', fgColor='0070C0')
                    cell.font = Font(name='Consolas', size=11, color='000000', bold=True)

            elif cell.row != 1:
                # 设置红球格式
                if cell.col_idx >= 3 and cell.col_idx <= 35:
                    if 'y' in str(cell.value):
                        cell.value = str(cell.value).replace('y', '')
                        cell.fill = PatternFill(fill_type='solid', fgColor='f2f1dd')
                        cell.font = Font(name='Consolas', size=11, color='D9D9D9', bold=True)
                    else:
                        cell.fill = PatternFill(fill_type='solid', fgColor='f2f1dd')
                        cell.font = Font(name='Consolas', size=11, color='F11515', bold=True)
                # 设置篮球格式
                elif cell.col_idx >= 36 and cell.col_idx <= 51:
                    if 'y' in str(cell.value):
                        cell.value = str(cell.value).replace('y', '')
                        cell.fill = PatternFill(fill_type='solid', fgColor='d9effc')
                        cell.font = Font(name='Consolas', size=11, color='D9D9D9', bold=True)
                    else:
                        cell.fill = PatternFill(fill_type='solid', fgColor='d9effc')
                        cell.font = Font(name='Consolas', size=11, color='1a428a', bold=True)
                # 设置连号次数格式
                else:
                    cell.font = Font(name='Consolas', size=11, color='000000', bold=True)

    # 保存文件
    wb.save('./Data/output.xlsx')


def check_file_exists(path):
    if os.path.exists(path=path):
        return True
    else:
        return False
